"""
FastAPI application for the Sales Forecasting system.

Endpoints
---------
GET  /health                  → liveness check
POST /pipeline/run            → trigger the full pipeline (background)
GET  /pipeline/status         → poll pipeline progress
GET  /data/summary            → workbook metadata
GET  /stores                  → list of store IDs
GET  /categories              → list of category codes
GET  /forecasts/{grain}       → Plotly-ready forecast JSON
GET  /forecasts/{grain}/export → download forecast XLSX
GET  /evaluation/metrics      → model comparison table
POST /upload                  → upload a new workbook

Design notes
------------
- Pipeline runs in a thread pool executor to avoid blocking the event loop.
- All pipeline state is in-memory (single-user local demo).
- Forecast artifacts are read from FORECAST_DIR / DATA_DIR after the pipeline
  completes — endpoints return 404 if the pipeline has not run yet.
- CORS is open (allow_origins=["*"]) for the local demo.
  Tighten this before any deployment.
"""

from __future__ import annotations

import asyncio
import io
import json
import logging
import shutil
import tempfile
import traceback
import uuid
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any

import pandas as pd
from fastapi import BackgroundTasks, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.responses import Response as StarletteResponse

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from config import (
    API_HOST, API_PORT,
    DATA_DIR, FORECAST_DIR, REPORT_DIR,
    WORKBOOK_PATH, WORKBOOK_PATH_RAW,
    SHEET_HEADER, SHEET_SALE_LINES, SHEET_ITEM_MASTER,
    IMPORTED_FLAT_PATH,
    ALLOWED_DB_TYPES, DB_QUERY_TIMEOUT, DB_MAX_ROWS,
    API_FETCH_TIMEOUT, API_MAX_ROWS,
)

logger = logging.getLogger(__name__)

app = FastAPI(
    title="Sales Forecasting API",
    version="1.0.0",
    description="Retail sales forecasting pipeline built on AI Forecasting Data.xlsx",
)

# ── CORS ──────────────────────────────────────────────────────────────────────
# Allow all origins for local development.  Restrict in production.
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve the UI from /ui path.
# Preferred location after repo split: <repo>/frontend
# Fallback for legacy layout: <backend>/ui
_repo_root = Path(__file__).resolve().parents[2]
_ui_candidates = [
    _repo_root / "frontend",
    Path(__file__).resolve().parent.parent / "ui",
]
for _ui_dir in _ui_candidates:
    if _ui_dir.exists():
        app.mount("/ui", StaticFiles(directory=str(_ui_dir), html=True), name="ui")
        break

# ── Pipeline state (in-memory, single-user) ───────────────────────────────────
_state: dict[str, Any] = {
    "status": "idle",        # idle | running | complete | error
    "job_id": None,
    "started_at": None,
    "finished_at": None,
    "error": None,
    "last_run_artifacts": {},
    "summary": None,
}

_executor = ThreadPoolExecutor(max_workers=1)

# ── Data prep state (in-memory, single-user) ─────────────────────────────────
_prep: dict[str, Any] = {
    "df": None,           # current DataFrame being prepped
    "history": [],        # list of previous df states for undo
    "source": None,       # "excel" or "flat_csv"
    "max_undo": 20,
}


# ── pipeline runner (sync, runs in executor) ──────────────────────────────────

def _run_pipeline_sync() -> None:
    """
    Full synchronous pipeline: ingest → clean → feature eng → train → forecast.
    Called via run_in_executor so it does not block the event loop.
    """
    import datetime

    _state["started_at"] = datetime.datetime.utcnow().isoformat()
    _state["error"] = None
    _state["last_run_artifacts"] = {}

    try:
        # Import here to avoid circular issues at module load time
        from data_ingestion.loader import load_workbook, get_workbook_summary
        from data_cleaning.cleaner import (
            clean_sale_lines, clean_header, clean_item_master,
            build_fact_sales, aggregate_to_daily, save_cleaned_data,
        )
        from data_cleaning.quarantine import QuarantineLog
        from models.baselines import get_all_baselines
        from models.ridge_model import RidgeForecaster
        from models.lasso_elnet_model import LassoForecaster, ElasticNetForecaster
        from models.forecaster import ForecastingPipeline
        from evaluation.metrics import compute_all_metrics, save_evaluation_report

        # 1. Load data (flat import takes priority if newer)
        logger.info("[pipeline] Step 1: Loading data …")
        use_flat = (
            IMPORTED_FLAT_PATH.exists()
            and (
                not WORKBOOK_PATH.exists()
                or IMPORTED_FLAT_PATH.stat().st_mtime > WORKBOOK_PATH.stat().st_mtime
            )
        )
        if use_flat:
            from data_ingestion.loader import load_flat_data
            logger.info("[pipeline] Using imported flat data: %s", IMPORTED_FLAT_PATH)
            sheets = load_flat_data(IMPORTED_FLAT_PATH, fmt="csv")
        else:
            sheets = load_workbook()
        summary = get_workbook_summary(sheets)

        # 2. Clean
        logger.info("[pipeline] Step 2: Cleaning data …")
        q = QuarantineLog()
        sl_clean  = clean_sale_lines(sheets["sale_lines"], q)
        hdr_clean = clean_header(sheets["header"], q)
        im_clean  = clean_item_master(sheets["item_master"], q)

        q.flush(DATA_DIR / "quarantine.csv")
        summary["quarantine"] = q.summary()

        # 3. Build fact table + aggregates
        logger.info("[pipeline] Step 3: Building canonical tables …")
        fact = build_fact_sales(sl_clean, hdr_clean, im_clean)
        fs_daily, item_daily, store_daily, cat_daily, total_daily = aggregate_to_daily(fact)
        save_cleaned_data(fs_daily, item_daily, store_daily, cat_daily, total_daily)

        # Save summary with data shape info
        summary["fact_rows"] = len(fs_daily)
        summary["item_daily_rows"] = len(item_daily)
        summary["store_daily_rows"] = len(store_daily)
        summary["category_daily_rows"] = len(cat_daily)
        summary["total_daily_rows"] = len(total_daily)
        _save_json(summary, REPORT_DIR / "summary.json")
        _state["summary"] = summary

        # 4+5. Train models + walk-forward + forecast for each grain
        logger.info("[pipeline] Step 4–5: Training and forecasting …")
        HORIZONS = [7, 14, 30]
        all_metrics: list[dict] = []

        models_to_run = {
            "baselines": get_all_baselines(),
            "ridge": RidgeForecaster(),
            "lasso": LassoForecaster(),
            "elnet": ElasticNetForecaster(),
        }

        grains = {
            "total":    (total_daily, [],           "total",     "sales_value"),
            "store":    (store_daily, ["store_id"], "store_id",  "sales_value"),
            "category": (cat_daily,  ["category"],  "category", "sales_value"),
            "item":     (item_daily, ["item_id"],   "item_id",  "sales_qty"),
        }

        FORECAST_DIR.mkdir(parents=True, exist_ok=True)

        for grain_name, (grain_df, group_cols, group_id_col, target_col) in grains.items():
            # For item grain, filter to items with enough history
            if grain_name == "item":
                from config import MIN_ITEM_DAYS
                item_day_counts = grain_df.groupby("item_id")["date"].nunique()
                eligible_items = item_day_counts[item_day_counts >= MIN_ITEM_DAYS].index.tolist()
                grain_df = grain_df[grain_df["item_id"].isin(eligible_items)].copy()
                logger.info(
                    "[pipeline] Item grain: %d items have >= %d days of data (out of %d total)",
                    len(eligible_items), MIN_ITEM_DAYS, len(item_day_counts),
                )

            groups = (
                grain_df[group_id_col].unique().tolist()
                if group_id_col in grain_df.columns
                else ["total"]
            )

            for model_tag, model_obj in models_to_run.items():
                if isinstance(model_obj, dict):
                    model_instances = model_obj   # baselines dict
                else:
                    # Use class name (e.g. "RidgeForecaster") not tag ("ridge")
                    # so file names are consistent with frontend dropdown values
                    model_instances = {model_obj.get_model_name(): model_obj}

                for m_name, m_instance in model_instances.items():
                    all_fc_records = []
                    for grp_val in groups:
                        for horizon in HORIZONS:
                            pipeline = ForecastingPipeline(
                                model=m_instance,
                                grain=grain_name,
                                group_cols=group_cols,
                                target_col=target_col,
                            )
                            try:
                                result = pipeline.run(
                                    grain_df,
                                    horizon=horizon,
                                    group_value=grp_val if group_id_col in grain_df.columns else None,
                                    add_uncertainty=(model_tag != "baselines"),
                                    n_bootstrap=20,  # reduced for speed in API context
                                )
                                # Collect forecast rows
                                fc_rows = result.forecast_df.copy()
                                fc_rows["grain"] = grain_name
                                fc_rows["group_key"] = grp_val
                                fc_rows["horizon"] = horizon
                                fc_rows["model_name"] = m_name
                                fc_rows["target_col"] = target_col

                                # For item grain, attach UOM and description metadata
                                if grain_name == "item":
                                    item_meta = grain_df[grain_df["item_id"] == grp_val].iloc[0]
                                    fc_rows["uom"] = item_meta.get("uom", "")
                                    fc_rows["description"] = item_meta.get("description", "")
                                    fc_rows["category"] = item_meta.get("category", "")

                                all_fc_records.append(fc_rows)

                                # Collect metrics
                                if result.validation:
                                    all_metrics.append(result.validation.to_dict())

                            except Exception as exc:
                                logger.warning(
                                    "Pipeline failed for grain=%s group=%s model=%s h=%d: %s",
                                    grain_name, grp_val, m_name, horizon, exc,
                                )

                    if all_fc_records:
                        fc_df = pd.concat(all_fc_records, ignore_index=True)
                        fc_path = FORECAST_DIR / f"forecast_{grain_name}_{m_name}.csv"
                        fc_df.to_csv(fc_path, index=False)
                        _state["last_run_artifacts"][f"forecast_{grain_name}_{m_name}"] = str(fc_path)

        # 6. Save evaluation report
        if all_metrics:
            save_evaluation_report(all_metrics, grain="all")

        _state["status"] = "complete"
        _state["finished_at"] = datetime.datetime.utcnow().isoformat()
        logger.info("[pipeline] Complete.")

    except Exception as exc:
        _state["status"] = "error"
        _state["error"] = traceback.format_exc()
        logger.error("[pipeline] FAILED: %s", exc, exc_info=True)


# ── helpers ───────────────────────────────────────────────────────────────────

def _save_json(data: Any, path: Path) -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    with open(path, "w") as f:
        json.dump(data, f, indent=2, default=str)


def _load_csv_or_404(path: Path) -> pd.DataFrame:
    """Load a CSV artifact; raise 404 if not yet created by the pipeline."""
    if not path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Artifact not found: {path.name}. Run the pipeline first.",
        )
    return pd.read_csv(path, parse_dates=["date"] if "date" in pd.read_csv(path, nrows=0).columns else [])


# ── endpoints ─────────────────────────────────────────────────────────────────

@app.get("/health")
async def health() -> dict:
    return {
        "status": "ok",
        "version": "1.0.0",
        "pipeline_status": _state["status"],
    }


@app.post("/pipeline/run")
async def run_pipeline() -> dict:
    """Trigger the full pipeline asynchronously."""
    if _state["status"] == "running":
        return {"status": "already_running", "job_id": _state["job_id"]}

    job_id = str(uuid.uuid4())
    _state["status"] = "running"
    _state["job_id"] = job_id

    loop = asyncio.get_event_loop()
    loop.run_in_executor(_executor, _run_pipeline_sync)

    return {"status": "started", "job_id": job_id}


@app.get("/pipeline/status")
async def pipeline_status() -> dict:
    return {
        "status": _state["status"],
        "job_id": _state["job_id"],
        "started_at":  _state["started_at"],
        "finished_at": _state["finished_at"],
        "error": _state["error"],
    }


@app.get("/data/summary")
async def data_summary() -> dict:
    """Return workbook metadata. 404 if pipeline has not run yet."""
    # Try in-memory first
    if _state["summary"]:
        return _state["summary"]

    summary_path = REPORT_DIR / "summary.json"
    if summary_path.exists():
        with open(summary_path) as f:
            return json.load(f)

    raise HTTPException(
        status_code=404,
        detail="Summary not available. Run the pipeline first.",
    )


@app.get("/stores")
async def list_stores() -> list[str]:
    path = DATA_DIR / "fact_store_daily.csv"
    df = _load_csv_or_404(path)
    return sorted(df["store_id"].dropna().unique().tolist())


@app.get("/items")
async def list_items() -> list[dict]:
    """
    Return all items with their UOM, description, category, and recent sales stats.
    Used by the Production Planner to show what items are available for forecasting.
    """
    path = DATA_DIR / "fact_item_daily.csv"
    df = _load_csv_or_404(path)

    # Aggregate per item: total qty sold, avg daily qty, UOM, description
    items = df.groupby("item_id", as_index=False).agg(
        description=("description", "first") if "description" in df.columns else ("item_id", "first"),
        uom=("uom", "first"),
        category=("category", "first"),
        total_qty=("sales_qty", "sum"),
        total_value=("sales_value", "sum"),
        days_with_sales=("date", "nunique"),
        avg_daily_qty=("sales_qty", "mean"),
    )
    items = items.sort_values("total_value", ascending=False)

    return items.to_dict(orient="records")


@app.get("/items/forecast_summary")
async def item_forecast_summary(
    horizon: int = Query(7, ge=7, le=30),
    model: str = Query("RidgeForecaster"),
) -> dict:
    """
    Return a production planning summary: for each item, the forecasted
    daily qty (NOS pieces or KGS kilos) over the given horizon.
    This is the core endpoint for sweet shop production decisions.
    """
    fc_path = FORECAST_DIR / f"forecast_item_{model}.csv"
    if not fc_path.exists():
        available = list(FORECAST_DIR.glob("forecast_item_*.csv"))
        if not available:
            raise HTTPException(404, "No item-level forecasts found. Run the pipeline first.")
        fc_path = available[0]
        model = fc_path.stem.replace("forecast_item_", "")

    df = pd.read_csv(fc_path, parse_dates=["date"])
    df = df[df["horizon"] == horizon]
    if "group_key" in df.columns:
        df["group_key"] = df["group_key"].astype(str)

    if df.empty:
        raise HTTPException(404, f"No item forecasts for horizon={horizon}.")

    # Build per-item summary
    items = []
    for item_id, grp in df.groupby("group_key"):
        grp = grp.sort_values("date")
        uom = grp["uom"].iloc[0] if "uom" in grp.columns else ""
        desc = grp["description"].iloc[0] if "description" in grp.columns else ""
        category = grp["category"].iloc[0] if "category" in grp.columns else ""
        total_forecast_qty = round(grp["forecast"].sum(), 2)
        avg_daily_qty = round(grp["forecast"].mean(), 2)

        items.append({
            "item_id": item_id,
            "description": desc,
            "uom": uom,
            "category": category,
            "total_forecast_qty": total_forecast_qty,
            "avg_daily_qty": avg_daily_qty,
            "horizon": horizon,
            "model": model,
            "forecast_dates": grp["date"].astype(str).tolist(),
            "forecast_values": grp["forecast"].round(2).tolist(),
            "ci_lower": grp["ci_lower"].round(2).tolist() if "ci_lower" in grp.columns else [],
            "ci_upper": grp["ci_upper"].round(2).tolist() if "ci_upper" in grp.columns else [],
        })

    # Sort by total forecasted quantity descending
    items.sort(key=lambda x: x["total_forecast_qty"], reverse=True)

    # Separate into NOS and KGS groups for the production planner
    nos_items = [i for i in items if i["uom"] == "NOS"]
    kgs_items = [i for i in items if i["uom"] == "KGS"]
    other_items = [i for i in items if i["uom"] not in ("NOS", "KGS")]

    return {
        "model": model,
        "horizon": horizon,
        "total_items": len(items),
        "nos_items": nos_items,
        "nos_count": len(nos_items),
        "kgs_items": kgs_items,
        "kgs_count": len(kgs_items),
        "other_items": other_items,
        "all_items": items,
    }


@app.get("/categories")
async def list_categories() -> list[str]:
    path = DATA_DIR / "fact_category_daily.csv"
    df = _load_csv_or_404(path)
    return sorted(df["category"].dropna().unique().tolist())


def _read_fc_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, parse_dates=["date"])


def _read_act_csv(path: Path) -> pd.DataFrame:
    return pd.read_csv(path, parse_dates=["date"])


@app.get("/forecasts/{grain}")
async def get_forecasts(
    grain: str,
    horizon: int = Query(7, ge=7, le=30),
    store_id: str | None = Query(None),
    category: str | None = Query(None),
    item_id: str | None = Query(None),
    model: str = Query("RidgeForecaster"),
) -> dict:
    """
    Return forecast + actuals for a given grain, horizon, and group.

    Response shape (Plotly-ready):
    {
      "dates": [...],
      "actuals": [...],
      "forecast": [...],
      "ci_lower": [...],
      "ci_upper": [...],
      "model_name": "...",
      "group_key": "...",
      "grain": "...",
      "horizon": 7,
    }
    """
    if grain not in ("total", "store", "category", "item"):
        raise HTTPException(400, f"Invalid grain '{grain}'. Use: total|store|category|item.")

    # Find forecast file
    fc_path = FORECAST_DIR / f"forecast_{grain}_{model}.csv"
    if not fc_path.exists():
        available = list(FORECAST_DIR.glob(f"forecast_{grain}_*.csv"))
        if not available:
            raise HTTPException(
                404, f"No forecasts found for grain='{grain}'. Run the pipeline first."
            )
        fc_path = available[0]
        model = fc_path.stem.replace(f"forecast_{grain}_", "")

    df = _read_fc_csv(fc_path)
    df = df[df["horizon"] == horizon]
    # Ensure group_key is string for consistent comparison
    if "group_key" in df.columns:
        df["group_key"] = df["group_key"].astype(str)

    # Filter by group
    group_key = "total"
    if grain == "store" and store_id:
        df = df[df["group_key"] == str(store_id)]
        group_key = store_id
    elif grain == "category" and category:
        df = df[df["group_key"] == str(category)]
        group_key = category
    elif grain == "item" and item_id:
        df = df[df["group_key"] == str(item_id)]
        group_key = item_id
    elif "group_key" in df.columns:
        first = df["group_key"].iloc[0] if len(df) > 0 else "total"
        df = df[df["group_key"] == first]
        group_key = first

    if df.empty:
        raise HTTPException(404, f"No forecast data for the specified filters.")

    df = df.sort_values("date")

    # Load actuals for context
    actuals_map = {
        "total":    (DATA_DIR / "fact_total_daily.csv",    "sales_value"),
        "store":    (DATA_DIR / "fact_store_daily.csv",    "sales_value"),
        "category": (DATA_DIR / "fact_category_daily.csv", "sales_value"),
        "item":     (DATA_DIR / "fact_item_daily.csv",     "sales_qty"),
    }
    actuals_df = pd.DataFrame()
    act_path, act_col = actuals_map.get(grain, (None, "sales_value"))
    if act_path and act_path.exists():
        act = _read_act_csv(act_path)
        if grain == "store" and store_id and "store_id" in act.columns:
            act = act[act["store_id"] == store_id]
        elif grain == "category" and category and "category" in act.columns:
            act = act[act["category"] == category]
        elif grain == "item" and item_id and "item_id" in act.columns:
            act = act[act["item_id"] == item_id]
        actuals_df = act[["date", act_col]].rename(
            columns={act_col: "actual"}
        ).sort_values("date")

    # For item grain, include UOM and description metadata
    item_meta = {}
    if grain == "item" and not df.empty:
        item_meta = {
            "uom": df["uom"].iloc[0] if "uom" in df.columns else "",
            "description": df["description"].iloc[0] if "description" in df.columns else "",
            "category": df["category"].iloc[0] if "category" in df.columns else "",
            "target_col": "sales_qty",
        }

    return {
        "grain": grain,
        "group_key": group_key,
        "model_name": model,
        "horizon": horizon,
        "dates": df["date"].astype(str).tolist(),
        "forecast": df["forecast"].round(2).tolist(),
        "ci_lower": df["ci_lower"].round(2).tolist() if "ci_lower" in df.columns else [],
        "ci_upper": df["ci_upper"].round(2).tolist() if "ci_upper" in df.columns else [],
        "actuals_dates": actuals_df["date"].astype(str).tolist() if not actuals_df.empty else [],
        "actuals": actuals_df["actual"].round(2).tolist() if not actuals_df.empty else [],
        **item_meta,
    }


@app.get("/forecasts/{grain}/export")
async def export_forecasts(
    grain: str,
    horizon: int = Query(7, ge=7, le=30),
    model: str = Query("RidgeForecaster"),
) -> StreamingResponse:
    """Download forecast as XLSX."""
    fc_path = FORECAST_DIR / f"forecast_{grain}_{model}.csv"
    if not fc_path.exists():
        available = list(FORECAST_DIR.glob(f"forecast_{grain}_*.csv"))
        if not available:
            raise HTTPException(404, f"No forecasts for grain='{grain}'.")
        fc_path = available[0]

    df = _read_fc_csv(fc_path)
    df = df[df["horizon"] == horizon].sort_values(["group_key", "date"])

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        df.to_excel(writer, sheet_name="Forecasts", index=False)

    buf.seek(0)
    filename = f"forecast_{grain}_{model}_h{horizon}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/evaluation/metrics")
async def get_metrics(
    grain: str = Query("all"),
    model: str | None = Query(None),
) -> dict:
    """Return evaluation metrics. Optionally filter by model name."""
    metrics_path = REPORT_DIR / f"metrics_{grain}.json"
    if not metrics_path.exists():
        metrics_path = REPORT_DIR / "metrics_all.json"
    if not metrics_path.exists():
        raise HTTPException(404, "Metrics not found. Run the pipeline first.")

    with open(metrics_path) as f:
        records = json.load(f)

    if model:
        records = [r for r in records if r.get("model_name") == model]

    # Sort by mean_mae ascending
    records = sorted(records, key=lambda r: r.get("mean_mae", float("inf")))
    return {"records": records, "count": len(records)}


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)) -> dict:
    """
    Accept Excel (.xlsx/.xls), CSV, JSON, or XML uploads.

    Excel files are validated for the three required sheets.
    Flat files (CSV/JSON/XML) are validated against the column schema.
    """
    fname = file.filename or ""
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""

    if ext not in ("xlsx", "xls", "csv", "json", "xml"):
        raise HTTPException(400, f"Unsupported file type '.{ext}'. Accepted: .xlsx, .xls, .csv, .json, .xml")

    # Write to temp file
    suffix = f".{ext}"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        contents = await file.read()
        tmp.write(contents)
        tmp_path = Path(tmp.name)

    try:
        if ext in ("xlsx", "xls"):
            # Excel path: validate sheets
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            required_sheets = [SHEET_HEADER, SHEET_SALE_LINES, SHEET_ITEM_MASTER]
            missing = [s for s in required_sheets if s not in wb.sheetnames]
            wb.close()
            if missing:
                tmp_path.unlink(missing_ok=True)
                raise HTTPException(
                    400,
                    f"Uploaded file is missing required sheets: {missing}. "
                    f"Found: {wb.sheetnames}",
                )
            # Read Sale Lines sheet to get row count + columns for UI
            try:
                sl_df = pd.read_excel(tmp_path, sheet_name=SHEET_SALE_LINES, nrows=5)
                sl_row_count = pd.read_excel(tmp_path, sheet_name=SHEET_SALE_LINES).shape[0]
                sl_columns = list(sl_df.columns)
                hdr_count = pd.read_excel(tmp_path, sheet_name=SHEET_HEADER).shape[0]
                im_count = pd.read_excel(tmp_path, sheet_name=SHEET_ITEM_MASTER).shape[0]
            except Exception:
                sl_row_count, sl_columns, hdr_count, im_count = 0, [], 0, 0
            # Save as workbook
            WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(tmp_path), str(WORKBOOK_PATH))
            return {
                "status": "uploaded",
                "source_type": "excel",
                "filename": fname,
                "saved_to": str(WORKBOOK_PATH),
                "row_count": sl_row_count,
                "columns": sl_columns,
                "sheets": {
                    SHEET_SALE_LINES: sl_row_count,
                    SHEET_HEADER: hdr_count,
                    SHEET_ITEM_MASTER: im_count,
                },
                "message": f"Excel file uploaded ({sl_row_count:,} transactions, {hdr_count:,} receipts, {im_count:,} items).",
            }
        else:
            # Flat file path: CSV / JSON / XML
            from data_ingestion.loader import validate_flat_schema
            fmt_map = {"csv": "csv", "json": "json", "xml": "xml"}
            fmt = fmt_map[ext]

            if fmt == "csv":
                df = pd.read_csv(tmp_path)
            elif fmt == "json":
                df = pd.read_json(tmp_path)
            elif fmt == "xml":
                df = pd.read_xml(tmp_path)

            validation = validate_flat_schema(df)
            if not validation["valid"]:
                tmp_path.unlink(missing_ok=True)
                raise HTTPException(
                    400,
                    f"Schema validation failed. Missing columns: {validation['missing']}. "
                    f"Found columns: {validation['columns']}",
                )

            # Save as canonical flat CSV
            IMPORTED_FLAT_PATH.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(IMPORTED_FLAT_PATH, index=False)
            tmp_path.unlink(missing_ok=True)

            return {
                "status": "uploaded",
                "source_type": fmt,
                "filename": fname,
                "saved_to": str(IMPORTED_FLAT_PATH),
                "row_count": validation["row_count"],
                "columns": validation["columns"],
                "sample_rows": validation["sample_rows"],
                "message": f"{fmt.upper()} file uploaded ({validation['row_count']} rows). Call POST /pipeline/run to process it.",
            }

    except HTTPException:
        raise
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(400, f"Could not read uploaded file: {exc}")


@app.get("/import/status")
async def import_status() -> dict:
    """Check if data has been imported and what state the pipeline is in."""
    has_workbook = WORKBOOK_PATH.exists() or WORKBOOK_PATH_RAW.exists()
    has_flat = IMPORTED_FLAT_PATH.exists()

    source_type = None
    filename = None
    row_count = 0
    columns: list[str] = []

    if has_flat:
        source_type = "flat_csv"
        filename = IMPORTED_FLAT_PATH.name
        try:
            df_peek = pd.read_csv(IMPORTED_FLAT_PATH, nrows=5)
            row_count = sum(1 for _ in open(IMPORTED_FLAT_PATH)) - 1  # minus header
            columns = list(df_peek.columns)
        except Exception:
            pass
    elif has_workbook:
        source_type = "excel"
        filename = WORKBOOK_PATH.name if WORKBOOK_PATH.exists() else WORKBOOK_PATH_RAW.name

    # Check if pipeline has produced artifacts
    pipeline_complete = (REPORT_DIR / "summary.json").exists()

    return {
        "has_data": has_workbook or has_flat,
        "source_type": source_type,
        "filename": filename,
        "row_count": row_count,
        "columns": columns,
        "pipeline_status": _state["status"],
        "pipeline_complete": pipeline_complete,
    }


@app.post("/connect/api")
async def connect_api(body: dict) -> dict:
    """
    Fetch data from a REST API / OData endpoint.

    Body:
      url         : str (required)
      method      : str (GET or POST, default GET)
      headers     : dict (optional)
      auth_type   : str (none, bearer, api_key, basic)
      auth_value  : str (token or key value)
      json_path   : str (dot-separated path to records array, e.g. "data.records")
      preview_only: bool (default false — if true, return 5-row sample without saving)
    """
    import httpx

    url = body.get("url")
    if not url:
        raise HTTPException(400, "URL is required.")

    method = body.get("method", "GET").upper()
    headers = body.get("headers") or {}
    auth_type = body.get("auth_type", "none")
    auth_value = body.get("auth_value", "")
    json_path = body.get("json_path", "")
    preview_only = body.get("preview_only", False)

    # Build auth headers
    if auth_type == "bearer" and auth_value:
        headers["Authorization"] = f"Bearer {auth_value}"
    elif auth_type == "api_key" and auth_value:
        headers["X-API-Key"] = auth_value

    try:
        async with httpx.AsyncClient(timeout=API_FETCH_TIMEOUT) as client:
            if method == "POST":
                resp = await client.post(url, headers=headers)
            else:
                resp = await client.get(url, headers=headers)
            resp.raise_for_status()
            data = resp.json()
    except httpx.HTTPStatusError as exc:
        raise HTTPException(400, f"API returned HTTP {exc.response.status_code}: {exc.response.text[:500]}")
    except Exception as exc:
        raise HTTPException(400, f"Failed to fetch from API: {exc}")

    # Extract records via json_path
    if json_path:
        for key in json_path.split("."):
            if isinstance(data, dict) and key in data:
                data = data[key]
            else:
                raise HTTPException(400, f"JSON path '{json_path}' not found in response.")

    if isinstance(data, list):
        df = pd.DataFrame(data)
    elif isinstance(data, dict):
        df = pd.DataFrame([data])
    else:
        raise HTTPException(400, "API response could not be converted to tabular data.")

    if len(df) > API_MAX_ROWS:
        raise HTTPException(400, f"API returned {len(df)} rows, exceeding limit of {API_MAX_ROWS}.")

    # Validate schema
    from data_ingestion.loader import validate_flat_schema
    validation = validate_flat_schema(df)

    if preview_only:
        return {
            "status": "preview",
            "valid": validation["valid"],
            "row_count": validation["row_count"],
            "columns": validation["columns"],
            "missing": validation["missing"],
            "sample_rows": validation["sample_rows"],
        }

    if not validation["valid"]:
        raise HTTPException(
            400,
            f"Schema validation failed. Missing columns: {validation['missing']}. "
            f"Found: {validation['columns']}",
        )

    # Save
    IMPORTED_FLAT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(IMPORTED_FLAT_PATH, index=False)

    return {
        "status": "imported",
        "source_type": "api",
        "url": url,
        "row_count": validation["row_count"],
        "columns": validation["columns"],
        "sample_rows": validation["sample_rows"],
        "message": f"API data imported ({validation['row_count']} rows). Call POST /pipeline/run to process it.",
    }


@app.post("/connect/database")
async def connect_database(body: dict) -> dict:
    """
    Connect to a database and import data.

    Body:
      db_type  : str (postgresql, mysql, sqlite, mssql)
      host     : str
      port     : int
      database : str (or file path for sqlite)
      username : str
      password : str
      query    : str (SQL query to execute)
      test_only: bool (default false)
    """
    db_type = body.get("db_type", "").lower()
    if db_type not in ALLOWED_DB_TYPES:
        raise HTTPException(400, f"Unsupported database type '{db_type}'. Allowed: {ALLOWED_DB_TYPES}")

    host = body.get("host", "localhost")
    port = body.get("port")
    database = body.get("database", "")
    username = body.get("username", "")
    password = body.get("password", "")
    query = body.get("query", "")
    test_only = body.get("test_only", False)

    if not query:
        raise HTTPException(400, "SQL query is required.")

    # Build SQLAlchemy connection URL
    try:
        from sqlalchemy import create_engine, text

        if db_type == "sqlite":
            url = f"sqlite:///{database}"
        elif db_type == "postgresql":
            port = port or 5432
            url = f"postgresql+psycopg2://{username}:{password}@{host}:{port}/{database}"
        elif db_type == "mysql":
            port = port or 3306
            url = f"mysql+pymysql://{username}:{password}@{host}:{port}/{database}"
        elif db_type == "mssql":
            port = port or 1433
            try:
                import pyodbc  # noqa: F401
                url = f"mssql+pyodbc://{username}:{password}@{host}:{port}/{database}?driver=ODBC+Driver+17+for+SQL+Server"
            except ImportError:
                raise HTTPException(
                    400,
                    "SQL Server connector requires pyodbc and ODBC Driver 17. "
                    "Install with: pip install pyodbc",
                )
        else:
            raise HTTPException(400, f"Unsupported db_type: {db_type}")

        engine = create_engine(url, connect_args={"connect_timeout": DB_QUERY_TIMEOUT} if db_type != "sqlite" else {})

        with engine.connect() as conn:
            result = conn.execute(text(query))
            rows = result.fetchmany(DB_MAX_ROWS)
            columns = list(result.keys())
            total_count = len(rows)

        df = pd.DataFrame(rows, columns=columns)

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(400, f"Database connection failed: {exc}")

    # Validate schema
    from data_ingestion.loader import validate_flat_schema
    validation = validate_flat_schema(df)

    if test_only:
        return {
            "status": "connected",
            "valid": validation["valid"],
            "db_type": db_type,
            "row_count": total_count,
            "columns": validation["columns"],
            "missing": validation["missing"],
            "sample_rows": validation["sample_rows"],
        }

    if not validation["valid"]:
        raise HTTPException(
            400,
            f"Schema validation failed. Missing columns: {validation['missing']}. "
            f"Found: {validation['columns']}",
        )

    # Save
    IMPORTED_FLAT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(IMPORTED_FLAT_PATH, index=False)

    return {
        "status": "imported",
        "source_type": "database",
        "db_type": db_type,
        "row_count": validation["row_count"],
        "columns": validation["columns"],
        "sample_rows": validation["sample_rows"],
        "message": f"Database data imported ({validation['row_count']} rows). Call POST /pipeline/run to process it.",
    }


# ── Data Prep endpoints ──────────────────────────────────────────────────────

def _load_prep_data() -> pd.DataFrame:
    """Load the imported data into _prep['df'] if not already loaded."""
    if _prep["df"] is not None:
        return _prep["df"]

    # Try flat import first (takes priority), then workbook
    if IMPORTED_FLAT_PATH.exists():
        _prep["df"] = pd.read_csv(IMPORTED_FLAT_PATH)
        _prep["source"] = "flat_csv"
    elif WORKBOOK_PATH.exists() or WORKBOOK_PATH_RAW.exists():
        from data_ingestion.loader import load_workbook
        path = WORKBOOK_PATH if WORKBOOK_PATH.exists() else WORKBOOK_PATH_RAW
        sheets = load_workbook(path)
        _prep["df"] = sheets["sale_lines"]
        _prep["source"] = "excel"
    else:
        raise HTTPException(404, "No data imported yet. Upload a file first.")

    _prep["history"] = []
    return _prep["df"]


def _column_profile(df: pd.DataFrame) -> list[dict]:
    """Build Tableau Prep-style column profiles."""
    profiles = []
    for col in df.columns:
        s = df[col]
        dtype_str = str(s.dtype)
        null_count = int(s.isna().sum())
        total = len(s)

        profile: dict[str, Any] = {
            "name": col,
            "dtype": dtype_str,
            "null_count": null_count,
            "null_pct": round(null_count / total * 100, 1) if total else 0,
            "unique_count": int(s.nunique()),
            "total_count": total,
        }

        # Numeric stats
        if pd.api.types.is_numeric_dtype(s):
            clean = s.dropna()
            profile["kind"] = "numeric"
            profile["min"] = round(float(clean.min()), 2) if len(clean) else None
            profile["max"] = round(float(clean.max()), 2) if len(clean) else None
            profile["mean"] = round(float(clean.mean()), 2) if len(clean) else None
            profile["median"] = round(float(clean.median()), 2) if len(clean) else None
            # Histogram bins (10 bins)
            if len(clean) > 0:
                import numpy as np
                counts, edges = np.histogram(clean, bins=min(10, len(clean.unique())))
                profile["histogram"] = {
                    "counts": counts.tolist(),
                    "edges": [round(float(e), 2) for e in edges],
                }
        elif pd.api.types.is_datetime64_any_dtype(s):
            clean = s.dropna()
            profile["kind"] = "datetime"
            profile["min"] = str(clean.min()) if len(clean) else None
            profile["max"] = str(clean.max()) if len(clean) else None
        else:
            profile["kind"] = "text"
            # Top values
            vc = s.value_counts().head(8)
            profile["top_values"] = [
                {"value": str(v), "count": int(c)} for v, c in vc.items()
            ]

        profiles.append(profile)
    return profiles


def _push_undo():
    """Save current df state for undo."""
    if _prep["df"] is not None:
        _prep["history"].append(_prep["df"].copy())
        if len(_prep["history"]) > _prep["max_undo"]:
            _prep["history"].pop(0)


@app.get("/data/preview")
async def data_preview(
    page: int = Query(1, ge=1),
    page_size: int = Query(100, ge=10, le=500),
    sort_col: str | None = Query(None),
    sort_asc: bool = Query(True),
) -> dict:
    """
    Return paginated data preview + column profiles.

    Used by the Prep view to display the editable data grid and
    column profile cards (Tableau Prep style).
    """
    df = _load_prep_data()

    # Sort if requested
    if sort_col and sort_col in df.columns:
        df = df.sort_values(sort_col, ascending=sort_asc, na_position="last")

    total_rows = len(df)
    total_pages = max(1, (total_rows + page_size - 1) // page_size)
    start = (page - 1) * page_size
    end = start + page_size
    page_df = df.iloc[start:end]

    return {
        "columns": list(df.columns),
        "rows": page_df.fillna("").to_dict(orient="records"),
        "row_indices": list(range(start, min(end, total_rows))),
        "total_rows": total_rows,
        "total_pages": total_pages,
        "page": page,
        "page_size": page_size,
        "profiles": _column_profile(df),
        "can_undo": len(_prep["history"]) > 0,
    }


@app.post("/data/transform")
async def data_transform(body: dict) -> dict:
    """
    Apply a data transformation and return the updated preview.

    Supported operations (body.operation):
      rename_column   : {column, new_name}
      remove_column   : {column}
      change_type     : {column, new_type}  (text, number, date)
      fill_nulls      : {column, strategy}  (zero, mean, median, mode, empty_string, custom)
                        + optional {value} for custom
      filter_rows     : {column, operator, value}
                        operators: eq, neq, gt, lt, gte, lte, contains, not_contains,
                                   starts_with, ends_with, is_null, is_not_null
      remove_duplicates: {columns}  (list of column names)
      replace_values  : {column, find, replace_with}
      sort            : {column, ascending}
      edit_cell       : {row_index, column, value}
      add_column      : {name, default_value}
      undo            : (no params — revert to previous state)
    """
    df = _load_prep_data()
    op = body.get("operation", "")

    if op == "undo":
        if _prep["history"]:
            _prep["df"] = _prep["history"].pop()
            return {"status": "ok", "operation": "undo", "message": "Reverted to previous state."}
        else:
            raise HTTPException(400, "Nothing to undo.")

    # Save undo state before mutating
    _push_undo()

    try:
        if op == "rename_column":
            old = body["column"]
            new = body["new_name"]
            if old not in df.columns:
                raise HTTPException(400, f"Column '{old}' not found.")
            if new in df.columns:
                raise HTTPException(400, f"Column '{new}' already exists.")
            df.rename(columns={old: new}, inplace=True)

        elif op == "remove_column":
            col = body["column"]
            if col not in df.columns:
                raise HTTPException(400, f"Column '{col}' not found.")
            df.drop(columns=[col], inplace=True)

        elif op == "change_type":
            col = body["column"]
            new_type = body["new_type"]
            if col not in df.columns:
                raise HTTPException(400, f"Column '{col}' not found.")
            if new_type == "text":
                df[col] = df[col].astype(str).replace("nan", "")
            elif new_type == "number":
                df[col] = pd.to_numeric(df[col], errors="coerce")
            elif new_type == "date":
                df[col] = pd.to_datetime(df[col], errors="coerce")
            else:
                raise HTTPException(400, f"Unknown type '{new_type}'. Use: text, number, date.")

        elif op == "fill_nulls":
            col = body["column"]
            strategy = body.get("strategy", "zero")
            if col not in df.columns:
                raise HTTPException(400, f"Column '{col}' not found.")
            if strategy == "zero":
                df[col] = df[col].fillna(0)
            elif strategy == "mean":
                df[col] = df[col].fillna(df[col].mean())
            elif strategy == "median":
                df[col] = df[col].fillna(df[col].median())
            elif strategy == "mode":
                mode_val = df[col].mode()
                df[col] = df[col].fillna(mode_val.iloc[0] if len(mode_val) else "")
            elif strategy == "empty_string":
                df[col] = df[col].fillna("")
            elif strategy == "custom":
                df[col] = df[col].fillna(body.get("value", ""))
            else:
                raise HTTPException(400, f"Unknown strategy '{strategy}'.")

        elif op == "filter_rows":
            col = body["column"]
            operator = body["operator"]
            val = body.get("value", "")
            if col not in df.columns:
                raise HTTPException(400, f"Column '{col}' not found.")
            s = df[col]
            if operator == "eq":
                mask = s.astype(str) == str(val)
            elif operator == "neq":
                mask = s.astype(str) != str(val)
            elif operator == "gt":
                mask = pd.to_numeric(s, errors="coerce") > float(val)
            elif operator == "lt":
                mask = pd.to_numeric(s, errors="coerce") < float(val)
            elif operator == "gte":
                mask = pd.to_numeric(s, errors="coerce") >= float(val)
            elif operator == "lte":
                mask = pd.to_numeric(s, errors="coerce") <= float(val)
            elif operator == "contains":
                mask = s.astype(str).str.contains(str(val), case=False, na=False)
            elif operator == "not_contains":
                mask = ~s.astype(str).str.contains(str(val), case=False, na=False)
            elif operator == "starts_with":
                mask = s.astype(str).str.startswith(str(val), na=False)
            elif operator == "ends_with":
                mask = s.astype(str).str.endswith(str(val), na=False)
            elif operator == "is_null":
                mask = s.isna() | (s.astype(str).str.strip() == "") | (s.astype(str) == "nan")
            elif operator == "is_not_null":
                mask = s.notna() & (s.astype(str).str.strip() != "") & (s.astype(str) != "nan")
            else:
                raise HTTPException(400, f"Unknown operator '{operator}'.")
            removed = int((~mask).sum())
            df = df[mask].reset_index(drop=True)

        elif op == "remove_duplicates":
            cols = body.get("columns", list(df.columns))
            before = len(df)
            df = df.drop_duplicates(subset=cols, keep="first").reset_index(drop=True)
            removed = before - len(df)

        elif op == "replace_values":
            col = body["column"]
            find_val = body["find"]
            replace_val = body["replace_with"]
            if col not in df.columns:
                raise HTTPException(400, f"Column '{col}' not found.")
            df[col] = df[col].astype(str).str.replace(str(find_val), str(replace_val), regex=False)

        elif op == "sort":
            col = body["column"]
            asc = body.get("ascending", True)
            if col not in df.columns:
                raise HTTPException(400, f"Column '{col}' not found.")
            df = df.sort_values(col, ascending=asc, na_position="last").reset_index(drop=True)

        elif op == "edit_cell":
            row_idx = body["row_index"]
            col = body["column"]
            val = body["value"]
            if col not in df.columns:
                raise HTTPException(400, f"Column '{col}' not found.")
            if row_idx < 0 or row_idx >= len(df):
                raise HTTPException(400, f"Row index {row_idx} out of range.")
            df.at[row_idx, col] = val

        elif op == "add_column":
            name = body["name"]
            default = body.get("default_value", "")
            if name in df.columns:
                raise HTTPException(400, f"Column '{name}' already exists.")
            df[name] = default

        else:
            # Pop the undo since nothing happened
            if _prep["history"]:
                _prep["history"].pop()
            raise HTTPException(400, f"Unknown operation '{op}'. See endpoint docs for supported ops.")

        _prep["df"] = df

    except HTTPException:
        raise
    except Exception as exc:
        # Revert on unexpected error
        if _prep["history"]:
            _prep["df"] = _prep["history"].pop()
        raise HTTPException(400, f"Transform failed: {exc}")

    return {
        "status": "ok",
        "operation": op,
        "total_rows": len(df),
        "total_columns": len(df.columns),
        "message": f"Applied '{op}' successfully.",
    }


@app.post("/data/save")
async def data_save() -> dict:
    """
    Persist the current prep state back to file.

    For Excel sources: saves to imported_flat.csv (pipeline will pick up the flat file).
    For flat sources: overwrites imported_flat.csv.
    """
    if _prep["df"] is None:
        raise HTTPException(400, "No data loaded in prep editor.")

    df = _prep["df"]
    IMPORTED_FLAT_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(IMPORTED_FLAT_PATH, index=False)

    return {
        "status": "saved",
        "row_count": len(df),
        "columns": list(df.columns),
        "saved_to": str(IMPORTED_FLAT_PATH),
        "message": f"Data saved ({len(df)} rows, {len(df.columns)} columns).",
    }


@app.post("/data/prep/reset")
async def data_prep_reset() -> dict:
    """Clear the in-memory prep state, forcing a reload from file on next access."""
    _prep["df"] = None
    _prep["history"] = []
    _prep["source"] = None
    return {"status": "reset", "message": "Prep state cleared."}


# ── Job Persistence endpoints ──────────────────────────────────────────────────

@app.post("/jobs/save")
async def save_current_job(job_name: str = "untitled") -> dict:
    """
    Save the current prep state (DataFrame) to the database for later resumption.

    Args:
        job_name: Name for this job (defaults to filename if not provided)

    Returns:
        dict with job_id and status
    """
    if _prep["df"] is None:
        raise HTTPException(400, "No data loaded to save as a job.")

    from job_persistence import save_job
    import uuid

    job_id = str(uuid.uuid4())[:8]
    filename = job_name or "job"
    df = _prep["df"]
    rows = len(df)
    cols = len(df.columns)

    success = save_job(
        job_id=job_id,
        filename=filename,
        stage="prepped",
        df=df,
        rows=rows,
        cols=cols,
        metadata={
            "source": _prep["source"],
            "columns": list(df.columns),
        }
    )

    if success:
        return {
            "status": "saved",
            "job_id": job_id,
            "filename": filename,
            "rows": rows,
            "cols": cols,
            "message": f"Job '{filename}' saved successfully.",
        }
    else:
        raise HTTPException(500, "Failed to save job to database.")


@app.get("/jobs/list")
async def list_saved_jobs() -> dict:
    """List all saved jobs."""
    from job_persistence import list_jobs

    jobs = list_jobs()
    return {
        "status": "success",
        "count": len(jobs),
        "jobs": jobs,
    }


@app.post("/jobs/load/{job_id}")
async def load_saved_job(job_id: str) -> dict:
    """
    Load a previously saved job from the database.

    Args:
        job_id: Job ID to load

    Returns:
        dict with status and loaded job details
    """
    from job_persistence import load_job

    job = load_job(job_id)
    if not job:
        raise HTTPException(404, f"Job '{job_id}' not found.")

    # Restore to prep state
    _prep["df"] = job["df"]
    _prep["source"] = job["metadata"].get("source", "unknown")

    return {
        "status": "loaded",
        "job_id": job_id,
        "filename": job["filename"],
        "stage": job["stage"],
        "rows": job["rows"],
        "cols": job["cols"],
        "message": f"Job '{job['filename']}' loaded successfully.",
    }


@app.delete("/jobs/{job_id}")
async def delete_saved_job(job_id: str) -> dict:
    """Delete a saved job."""
    from job_persistence import delete_job

    if delete_job(job_id):
        return {"status": "deleted", "job_id": job_id}
    else:
        raise HTTPException(500, "Failed to delete job.")


# ── LLM Chat endpoints ──────────────────────────────────────────────────────

@app.post("/chat")
async def chat_endpoint(body: dict):
    """
    Stream an LLM response via Server-Sent Events.

    Body:
      messages      : list[{role, content}]  — conversation history
      screen_context: dict (optional) — {currentView, grain, horizon, model, groupValue}
      persona       : str (optional)  — default | ceo | supply_chain | reasoning
      provider      : str (optional)  — override active_provider
      model         : str (optional)  — override active_model
    """
    from api.llm_gateway import stream_chat

    messages = body.get("messages", [])
    if not messages:
        raise HTTPException(400, "messages list is required.")

    screen_context = body.get("screen_context")
    persona = body.get("persona", "default")
    provider = body.get("provider")
    model = body.get("model")

    async def event_stream():
        async for chunk in stream_chat(
            messages=messages,
            app_state=_state,
            prep_state=_prep,
            screen_context=screen_context,
            persona=persona,
            provider_override=provider,
            model_override=model,
        ):
            yield chunk

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/llm/config")
async def llm_config_get() -> dict:
    """Return current LLM configuration (keys redacted)."""
    from api.llm_gateway import load_config

    cfg = load_config(force=True)
    # Redact API keys
    safe = json.loads(json.dumps(cfg, default=str))
    for prov in safe.get("providers", {}).values():
        key = prov.get("api_key", "")
        if key:
            prov["api_key"] = key[:4] + "..." + key[-4:] if len(key) > 8 else "***"
    return safe


@app.post("/llm/config")
async def llm_config_update(body: dict) -> dict:
    """
    Update LLM configuration.

    Body (all optional):
      active_provider: str
      active_model   : str
      api_key        : str  — sets key for active_provider
      provider       : str  — which provider's key to set (if different from active)
    """
    from api.llm_gateway import load_config, save_config

    cfg = load_config(force=True)

    if "active_provider" in body:
        cfg["active_provider"] = body["active_provider"]
    if "active_model" in body:
        cfg["active_model"] = body["active_model"]
    if "api_key" in body:
        prov = body.get("provider", cfg.get("active_provider", "gemini"))
        if prov in cfg.get("providers", {}):
            cfg["providers"][prov]["api_key"] = body["api_key"]

    save_config(cfg)
    return {"status": "updated", "active_provider": cfg["active_provider"], "active_model": cfg.get("active_model")}
