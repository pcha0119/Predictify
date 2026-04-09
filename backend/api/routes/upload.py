"""Routes: POST /upload, GET /import/status"""

from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, File, HTTPException, UploadFile

import sys
sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
from config import (
    WORKBOOK_PATH, WORKBOOK_PATH_RAW, IMPORTED_FLAT_PATH, REPORT_DIR,
    SHEET_HEADER, SHEET_SALE_LINES, SHEET_ITEM_MASTER,
)
from api.state import _state

router = APIRouter(tags=["upload"])


@router.post("/upload")
async def upload_file(file: UploadFile = File(...)) -> dict:
    """Accept Excel (.xlsx/.xls), CSV, JSON, or XML uploads."""
    fname = file.filename or ""
    ext = fname.rsplit(".", 1)[-1].lower() if "." in fname else ""

    if ext not in ("xlsx", "xls", "csv", "json", "xml"):
        raise HTTPException(400, f"Unsupported file type '.{ext}'. Accepted: .xlsx, .xls, .csv, .json, .xml")

    suffix = f".{ext}"
    with tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
        contents = await file.read()
        tmp.write(contents)
        tmp_path = Path(tmp.name)

    try:
        if ext in ("xlsx", "xls"):
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
            required_sheets = [SHEET_HEADER, SHEET_SALE_LINES, SHEET_ITEM_MASTER]
            missing = [s for s in required_sheets if s not in wb.sheetnames]
            wb.close()
            if missing:
                tmp_path.unlink(missing_ok=True)
                raise HTTPException(
                    400,
                    f"Uploaded file is missing required sheets: {missing}. Found: {wb.sheetnames}",
                )
            try:
                sl_df = pd.read_excel(tmp_path, sheet_name=SHEET_SALE_LINES, nrows=5)
                sl_row_count = pd.read_excel(tmp_path, sheet_name=SHEET_SALE_LINES).shape[0]
                sl_columns = list(sl_df.columns)
                hdr_count = pd.read_excel(tmp_path, sheet_name=SHEET_HEADER).shape[0]
                im_count = pd.read_excel(tmp_path, sheet_name=SHEET_ITEM_MASTER).shape[0]
            except Exception:
                sl_row_count, sl_columns, hdr_count, im_count = 0, [], 0, 0

            WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(tmp_path), str(WORKBOOK_PATH))
            return {
                "status": "uploaded",
                "source_type": "excel",
                "filename": fname,
                "saved_to": str(WORKBOOK_PATH),
                "row_count": sl_row_count,
                "columns": sl_columns,
                "sheets": {
                    SHEET_SALE_LINES: sl_row_count,
                    SHEET_HEADER: hdr_count,
                    SHEET_ITEM_MASTER: im_count,
                },
                "message": f"Excel file uploaded ({sl_row_count:,} transactions, {hdr_count:,} receipts, {im_count:,} items).",
            }
        else:
            from data_ingestion.loader import validate_flat_schema
            fmt_map = {"csv": "csv", "json": "json", "xml": "xml"}
            fmt = fmt_map[ext]

            if fmt == "csv":
                df = pd.read_csv(tmp_path)
            elif fmt == "json":
                df = pd.read_json(tmp_path)
            elif fmt == "xml":
                df = pd.read_xml(tmp_path)

            validation = validate_flat_schema(df)
            if not validation["valid"]:
                tmp_path.unlink(missing_ok=True)
                raise HTTPException(
                    400,
                    f"Schema validation failed. Missing columns: {validation['missing']}. "
                    f"Found columns: {validation['columns']}",
                )

            IMPORTED_FLAT_PATH.parent.mkdir(parents=True, exist_ok=True)
            df.to_csv(IMPORTED_FLAT_PATH, index=False)
            tmp_path.unlink(missing_ok=True)

            return {
                "status": "uploaded",
                "source_type": fmt,
                "filename": fname,
                "saved_to": str(IMPORTED_FLAT_PATH),
                "row_count": validation["row_count"],
                "columns": validation["columns"],
                "sample_rows": validation["sample_rows"],
                "message": f"{fmt.upper()} file uploaded ({validation['row_count']} rows). Call POST /pipeline/run to process it.",
            }

    except HTTPException:
        raise
    except Exception as exc:
        tmp_path.unlink(missing_ok=True)
        raise HTTPException(400, f"Could not read uploaded file: {exc}")


@router.get("/import/status")
async def import_status() -> dict:
    """Check if data has been imported and what state the pipeline is in."""
    has_workbook = WORKBOOK_PATH.exists() or WORKBOOK_PATH_RAW.exists()
    has_flat = IMPORTED_FLAT_PATH.exists()

    source_type = None
    filename = None
    row_count = 0
    columns: list[str] = []

    if has_flat:
        source_type = "flat_csv"
        filename = IMPORTED_FLAT_PATH.name
        try:
            df_peek = pd.read_csv(IMPORTED_FLAT_PATH, nrows=5)
            row_count = sum(1 for _ in open(IMPORTED_FLAT_PATH)) - 1
            columns = list(df_peek.columns)
        except Exception:
            pass
    elif has_workbook:
        source_type = "excel"
        filename = WORKBOOK_PATH.name if WORKBOOK_PATH.exists() else WORKBOOK_PATH_RAW.name

    pipeline_complete = (REPORT_DIR / "summary.json").exists()

    return {
        "has_data": has_workbook or has_flat,
        "source_type": source_type,
        "filename": filename,
        "row_count": row_count,
        "columns": columns,
        "pipeline_status": _state["status"],
        "pipeline_complete": pipeline_complete,
    }
